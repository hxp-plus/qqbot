import openpyxl
from openpyxl import load_workbook
import re
import sys

book_name = sys.argv[1]
datafile = 'data.txt'
offset_date = 20
offset_student_number = 3
row_start = 3

def read_xlsx(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return workbook,sheet

def write_xlsx(student_number, date, book_name, offset_student_number, offset_date):
    workbook = load_workbook(book_name)
    sheet = workbook.active
    for i in range(1,50):
        if sheet.cell(row=i, column=offset_student_number).value == student_number :
            for j in range(offset_date, offset_date + 3):
                if sheet.cell(row=i, column=j).value == None:
                    sheet.cell(row=i, column=j).value = date
                    workbook.save(book_name)
                    break


def read_data_and_write(datafile, book_name, offset_student_number, offset_date):
    fo = open(datafile, 'r+')
    while True:
        line = fo.readline()
        if not line:
            break
        if re.search(r'U\d{9}', line) and re.search(r'\d{4}/\d{1,2}/\d{1,2}', line):
            student_number=re.search(r'U\d{9}', line).group(0)
            date=re.search(r'\d{4}/\d{1,2}/\d{1,2}', line).group(0)
            write_xlsx(student_number, date, book_name, offset_student_number, offset_date)

def init(book_name, row_start):
    workbook = load_workbook(book_name)
    sheet = workbook.active
    for i in range(row_start,50):
        for j in range(offset_date, offset_date + 3):
            sheet.cell(row=i, column=j).value = ''

init(book_name, row_start)
read_data_and_write(datafile, book_name, offset_student_number, offset_date)
